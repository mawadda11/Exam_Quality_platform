"""Raw xlsx reading for the 11 approved KB workbooks. Deliberately dumb: no
validation or normalization here - just turns each workbook into a list of
RawRow. validator.py and normalizer.py both consume this same raw form.
"""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from app.services.knowledge_base.schemas import WorkbookSchema


class KnowledgeBaseFileMissingError(RuntimeError):
    """Raised when an expected workbook file is not present in source_dir."""


@dataclass(frozen=True)
class RawRow:
    workbook: str
    sheet: str | None
    row_number: int  # 1-based position within the sheet, header row = 1
    values: dict[str, object]


@dataclass(frozen=True)
class RawWorkbook:
    schema: WorkbookSchema
    sheet: str | None
    header: tuple[str, ...]
    rows: tuple[RawRow, ...]
    file_hash: str


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_workbook(source_dir: Path, schema: WorkbookSchema) -> RawWorkbook:
    path = source_dir / schema.filename
    if not path.is_file():
        raise KnowledgeBaseFileMissingError(f"Missing required workbook: {schema.filename}")

    file_hash = sha256_file(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active
    sheet_name = sheet.title if sheet is not None else None
    all_rows = list(sheet.iter_rows(values_only=True)) if sheet is not None else []
    if not all_rows:
        return RawWorkbook(schema=schema, sheet=sheet_name, header=(), rows=(), file_hash=file_hash)

    header = tuple("" if h is None else str(h) for h in all_rows[0])
    rows = tuple(
        RawRow(
            workbook=schema.filename,
            sheet=sheet_name,
            row_number=index,
            values=dict(zip(header, values, strict=False)),
        )
        for index, values in enumerate(all_rows[1:], start=2)
    )
    return RawWorkbook(
        schema=schema, sheet=sheet_name, header=header, rows=rows, file_hash=file_hash
    )
